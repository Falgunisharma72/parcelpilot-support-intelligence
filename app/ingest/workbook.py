"""Workbook -> SQLite.

Structured data goes into SQLite rather than staying as in-memory dataframes for
one reason that matters to this brief: access control. Scoping a query with
`WHERE account_id = ?` in the data layer is enforceable and auditable; filtering
a Python list *after* loading everything means the unfiltered rows existed in
the agent's process, one bug away from a leak.

It also gives us somewhere real for state-changing actions to land (escalations,
tasks, ticket updates) and a durable audit log.
"""
from __future__ import annotations

import sqlite3
from datetime import datetime
from pathlib import Path

import openpyxl

from app.config import BUILD_DIR, DEFAULT_SNAPSHOT, WORKBOOK, parse_ts

SCHEMA = """
CREATE TABLE IF NOT EXISTS accounts (
    account_id TEXT PRIMARY KEY, account_name TEXT, plan TEXT, status TEXT,
    csm TEXT, contract_file TEXT, premium_support INTEGER, notes TEXT
);
CREATE TABLE IF NOT EXISTS orders (
    order_id TEXT PRIMARY KEY, account_id TEXT, carrier TEXT, status TEXT,
    booked_at TEXT, pickup_window_start TEXT, pickup_window_end TEXT,
    pickup_actual_at TEXT, shipment_fee_inr REAL, carrier_fault INTEGER,
    customer_fault INTEGER, cancellation_requested_at TEXT, notes TEXT
);
CREATE TABLE IF NOT EXISTS tickets (
    ticket_id TEXT PRIMARY KEY, account_id TEXT, created_at TEXT, status TEXT,
    subject TEXT, description TEXT, channel TEXT, assigned_to TEXT,
    last_customer_message_at TEXT, historical_resolution TEXT
);
CREATE TABLE IF NOT EXISTS meta (key TEXT PRIMARY KEY, value TEXT);

-- Mutable state written by the agent's action tools. Mocked systems of record,
-- but real rows: an action either happened or it did not.
CREATE TABLE IF NOT EXISTS escalations (
    escalation_id TEXT PRIMARY KEY, account_id TEXT, ticket_id TEXT,
    severity TEXT, summary TEXT, justification TEXT, requested_action TEXT,
    created_by TEXT, created_by_role TEXT, created_at TEXT, status TEXT,
    proposal_id TEXT
);
CREATE TABLE IF NOT EXISTS tasks (
    task_id TEXT PRIMARY KEY, account_id TEXT, ticket_id TEXT, title TEXT,
    details TEXT, due_at TEXT, owner TEXT, created_by TEXT, created_at TEXT,
    status TEXT, proposal_id TEXT
);
CREATE TABLE IF NOT EXISTS ticket_updates (
    update_id INTEGER PRIMARY KEY AUTOINCREMENT, ticket_id TEXT, field TEXT,
    old_value TEXT, new_value TEXT, note TEXT, updated_by TEXT, updated_at TEXT,
    proposal_id TEXT
);
CREATE TABLE IF NOT EXISTS audit_log (
    id INTEGER PRIMARY KEY AUTOINCREMENT, ts TEXT, session_id TEXT,
    user_id TEXT, role TEXT, account_id TEXT, tool TEXT, args TEXT,
    outcome TEXT, detail TEXT
);
CREATE INDEX IF NOT EXISTS idx_orders_account ON orders(account_id);
CREATE INDEX IF NOT EXISTS idx_tickets_account ON tickets(account_id);
CREATE INDEX IF NOT EXISTS idx_audit_session ON audit_log(session_id);
"""

_TABLES = {
    "accounts": ["account_id", "account_name", "plan", "status", "csm",
                 "contract_file", "premium_support", "notes"],
    "orders": ["order_id", "account_id", "carrier", "status", "booked_at",
               "pickup_window_start", "pickup_window_end", "pickup_actual_at",
               "shipment_fee_inr", "carrier_fault", "customer_fault",
               "cancellation_requested_at", "notes"],
    "tickets": ["ticket_id", "account_id", "created_at", "status", "subject",
                "description", "channel", "assigned_to",
                "last_customer_message_at", "historical_resolution"],
}

_BOOL_COLS = {"premium_support", "carrier_fault", "customer_fault"}
_TS_COLS = {"booked_at", "pickup_window_start", "pickup_window_end",
            "pickup_actual_at", "cancellation_requested_at", "created_at",
            "last_customer_message_at"}


def _coerce(col: str, value):
    if value is None or value == "":
        return None
    if col in _BOOL_COLS:
        if isinstance(value, bool):
            return int(value)
        return int(str(value).strip().lower() in ("true", "1", "yes"))
    if col in _TS_COLS:
        dt = parse_ts(value)
        return dt.strftime("%Y-%m-%d %H:%M") if dt else None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    return value


def read_snapshot(workbook: Path = WORKBOOK) -> datetime:
    """Read the dataset snapshot declared in the README sheet.

    The brief is explicit that this, not wall-clock time, is the reference for
    every time-based question - so it is read from the workbook rather than
    copied into code.
    """
    wb = openpyxl.load_workbook(workbook, data_only=True, read_only=True)
    try:
        if "README" not in wb.sheetnames:
            return DEFAULT_SNAPSHOT
        for row in wb["README"].iter_rows(values_only=True):
            if row and row[0] and "snapshot" in str(row[0]).lower():
                raw = str(row[1])
                # e.g. "2026-08-16 11:00 Asia/Kolkata"
                stamp = " ".join(raw.split()[:2])
                dt = parse_ts(stamp)
                if dt:
                    return dt
        return DEFAULT_SNAPSHOT
    finally:
        wb.close()


def build_database(workbook: Path = WORKBOOK, db_path: Path | None = None,
                   force: bool = True) -> Path:
    db_path = db_path or (BUILD_DIR / "parcelpilot.db")
    db_path.parent.mkdir(parents=True, exist_ok=True)
    if force and db_path.exists():
        db_path.unlink()

    wb = openpyxl.load_workbook(workbook, data_only=True)
    conn = sqlite3.connect(db_path)
    try:
        conn.executescript(SCHEMA)
        for table, cols in _TABLES.items():
            if table not in wb.sheetnames:
                continue
            rows = list(wb[table].iter_rows(values_only=True))
            if not rows:
                continue
            header = [str(h).strip() for h in rows[0]]
            payload = []
            for raw in rows[1:]:
                record = dict(zip(header, raw))
                payload.append(tuple(_coerce(c, record.get(c)) for c in cols))
            placeholders = ",".join("?" * len(cols))
            conn.executemany(
                f"INSERT OR REPLACE INTO {table} ({','.join(cols)}) VALUES ({placeholders})",
                payload,
            )
        snapshot = read_snapshot(workbook)
        conn.execute("INSERT OR REPLACE INTO meta (key, value) VALUES (?,?)",
                     ("snapshot", snapshot.isoformat()))
        conn.execute("INSERT OR REPLACE INTO meta (key, value) VALUES (?,?)",
                     ("built_from", str(workbook.name)))
        conn.commit()
    finally:
        conn.close()
        wb.close()
    return db_path


if __name__ == "__main__":  # pragma: no cover
    path = build_database()
    print(f"built {path}")
    conn = sqlite3.connect(path)
    for t in ("accounts", "orders", "tickets"):
        print(t, conn.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0])
    print("snapshot", conn.execute("SELECT value FROM meta WHERE key='snapshot'").fetchone()[0])
